from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from terminaltables import AsciiTable

import os, csv, msoffcrypto, io

OUTPUTS_DIR = "./outputs"
# Contains the column mapping in .config
# The key is the cloudtalk CSV column name.
# The value is the column name(s) of the xl file to map to the key.
config_map:dict[str, str] = {}
wb:Workbook
xl_file_name:str

# Converts a single sheet of the xl file.
def convert_sheet_to_csv(sheet_name:str):
    print(f"Converting \"{sheet_name}\" sheet.")

    xl_name = xl_file_name.split(".")[0]
    sheet:Worksheet = wb[sheet_name]
    first_row:list[Cell] = sheet[1]

    # Returns a cell from the row at the given column name.
    def get_cell_from_column_name(column_name:str, row:list[Cell]) -> Cell:
        column_letter:str = ""
        for cell in first_row:
            if cell.value == column_name:
                column_letter = cell.column_letter
                break
        
        cells:list[Cell] = [c for c in row if c.column_letter == column_letter]

        if len(cells) == 0:
            raise Exception(f"The column \"{column_name}\" was not found in sheet \"{sheet_name}\" of document \"{xl_name}\".\nPlease change the column name or the config file to match each other.")
        
        return cells[0]

    
    with open(f"{OUTPUTS_DIR}/{xl_name}-{sheet_name}.csv", "w", newline = "") as file:
        writer = csv.writer(file, delimiter = ";")
        writer.writerow(config_map.keys())

        # Used for drawing a small table.
        first_two_rows:list[list] = []

        # Loop through each row
        for row in sheet.iter_rows(min_row = 2):
            data:list[str] = []

            # Loop through each mapping in the config.
            for mapping in config_map.values():
                if len(mapping) == 0 or mapping.isspace():
                    data.append("")
                    continue

                # Handle merging columns.
                elif "+" in mapping:
                    column_names:list[str] = mapping.split("+")
                    # Get the cells of each column.
                    cells:list[Cell] = [get_cell_from_column_name(name.strip(), row) for name in column_names]

                    combined_value:str = ""
                    for cell in cells:
                        combined_value += str(cell.value) + " "
                    
                    data.append(combined_value.strip())

                # Normal 1:1 mapping.
                else:
                    data.append(get_cell_from_column_name(mapping, row).value)

            if len(first_two_rows) < 2:
                first_two_rows.append(data)

            writer.writerow(data)

    # Display first two rows (with column names) of csv to user.
    first_two_rows.insert(0, config_map.keys())
    print(AsciiTable(first_two_rows).table)
    if sheet.max_row - 3 > 0:
        print(f"... {sheet.max_row - 3} more")


if __name__ == "__main__":
    try:
        if not os.path.exists(OUTPUTS_DIR):
            os.mkdir(OUTPUTS_DIR)

        config:str = open("./config.txt")
        for line in config:
            if line.startswith("//") or len(line) == 0 or line.isspace():
                continue

            line = line.strip()
            values:list[str] = line.split("=")
            config_map[values[0].strip()] = values[1].strip()
            
            if len(values[1]) == 0 or values[1].isspace():
                print(f"Warning: The column \"{values[0]}\" is unset.")

        xl_files:list[str] = [file for file in os.listdir(".") if ".xlsx" in file]

        if len(xl_files) == 0:
            raise Exception("No .xlsx files found in this directory.")
        
        for file in xl_files:
            xl_file_name = file

            # Try to open the xl file.
            # This will catch if the file is password locked.
            try:
                wb = load_workbook(xl_file_name)
            except:
                unlocked_wb = io.BytesIO()
                password:str = input(f"The document \"{xl_file_name}\" is locked by a password. Please enter it to continue: ").strip()
                
                while True:
                    try:
                        office_file = msoffcrypto.OfficeFile(open(xl_file_name, "rb"))
                        office_file.load_key(password = password)
                        office_file.decrypt(unlocked_wb)
                        break
                    except:
                        password = input("Incorrect password entered. Try again: ").strip()
                    
                wb = load_workbook(unlocked_wb)

            print(f"Converting \"{xl_file_name}\" document.")

            for sheet in wb.sheetnames:
                convert_sheet_to_csv(sheet)
            print()

        input("Finished!\nPress ENTER to close.")
    
    except Exception as e:
        input(f"\nError: {str(e)}\nPress ENTER to exit.")